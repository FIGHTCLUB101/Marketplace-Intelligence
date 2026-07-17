"""Combines per-shard scraper output files into a single, canonically
ordered output file. Used by parallel_runner.py so a parallelized scrape
still produces one file at the same path the rest of the pipeline reads."""
import os
from pathlib import Path

from openpyxl import Workbook, load_workbook


def merge_shards(shard_paths, output_path, columns, sort_key_fn):
    """Reads whatever rows exist across shard_paths (skipping any that
    don't exist yet -- a worker may not have saved its first row), sorts
    them with sort_key_fn, and atomically replaces output_path with the
    result. Returns the number of rows written. Raises PermissionError if
    output_path is locked by another program (e.g. open in Excel) --
    callers should treat that as "try again next cycle", not fatal."""
    rows = []
    for path in shard_paths:
        path = Path(path)
        if not path.exists():
            continue
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            wb.close()
            continue
        for row in rows_iter:
            rows.append(dict(zip(header, row)))
        wb.close()

    rows.sort(key=sort_key_fn)

    output_path = Path(output_path)
    tmp_path = output_path.with_name(output_path.stem + ".tmp" + output_path.suffix)
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.append(columns)
    for row in rows:
        out_ws.append([row.get(c) for c in columns])
    out_wb.save(tmp_path)
    out_wb.close()
    os.replace(tmp_path, output_path)  # atomic on the same filesystem
    return len(rows)
