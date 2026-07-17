"""Shared reliability toolkit for the 4 oats-brand scrapers.

Fixes the crash pattern diagnosed during Sprint 1 planning: Chrome's
renderer runs out of memory after ~200 localities of sustained DOM-heavy
automation (repeated full-page reloads, driver.page_source pulls, blind
modal enumeration in set_location), crashing mid-locality with no
recovery -- which cascades into every remaining locality failing the
same way. This module adds: periodic driver recycling before memory
pressure reaches that point, detection of a dead session so the current
locality can be retried instead of the whole run silently degrading,
real CAPTCHA/block detection with a pause-and-wait loop (all 4 scrapers
previously had inconsistent or, for Zepto, entirely absent block
handling), and incremental xlsx saves that don't get slower as the run
accumulates rows.
"""
import random
import time
from pathlib import Path

from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

BLOCK_TITLE_KEYWORDS = ["captcha", "robot", "blocked", "verify"]
BLOCK_BODY_MARKERS = [
    "access denied",
    "unusual traffic",
    "attention required",
    "awswafintegration",
    "challenge-container",
    "are you human",
    # Zepto's rate-limit-style wall ("Oops! Please login to continue
    # searching") -- confirmed live: it silently returned 0 results for two
    # searches in a row before clearing on its own a couple of searches
    # later. Matched on the specific phrase, not bare "login" -- every
    # normal Zepto page already has a "Login" link in the header.
    "please login to continue searching",
]

DEAD_SESSION_MARKERS = [
    "chrome not reachable",
    "invalid session id",
    "session deleted",
    "no such window",
    "disconnected",
]


def is_blocked(driver) -> bool:
    title = (driver.title or "").lower()
    if any(k in title for k in BLOCK_TITLE_KEYWORDS):
        return True
    body = (driver.page_source or "").lower()
    return any(m in body for m in BLOCK_BODY_MARKERS)


def wait_for_manual_unblock(driver, beep_fn, max_wait_s=180, poll_s=3) -> bool:
    if not is_blocked(driver):
        return True
    beep_fn()
    waited = 0.0
    while waited < max_wait_s:
        time.sleep(poll_s)
        waited += poll_s
        if not is_blocked(driver):
            return True
    return False


def keep_window_unminimized(driver) -> None:
    """The real, verified root cause of "scraping breaks when the window is
    minimized": a genuinely OS-minimized (SW_MINIMIZE) Chrome window has its
    rendering pipeline suspended for real content -- Zepto and Swiggy's own
    card lists depend on requestAnimationFrame/real layout to mount lazily
    loaded items, and Chrome pauses that for a truly minimized top-level
    window regardless of any page-JS trick. Confirmed live, with a clean
    before/after control on identical queries: minimized-before-navigation
    returned 0 real cards for both Zepto and Swiggy (document.body.innerText
    came back empty even though document.hidden reported False via
    defeat_visibility_throttling above -- so that fix alone doesn't reach
    this). A small window that stays in the normal/restored OS window state
    (never minimized) renders correctly regardless of its size -- confirmed
    live: shrinking to 350x280 in the corner instead of minimizing produced
    the full, real card list for both sites.

    This only solves the problem if nothing subsequently sends the window a
    real OS minimize afterward (e.g. clicking the taskbar icon, Win+Down) --
    that still suspends rendering no matter what, since it's Chrome's own
    power-saving behavior for a minimized top-level window, not a flag or
    script we can override. The practical instruction for whoever is running
    this: leave the window alone once it shrinks -- don't minimize it.
    """
    try:
        driver.set_window_rect(x=0, y=0, width=350, height=280)
    except Exception:
        pass


def defeat_visibility_throttling(driver) -> None:
    """Chrome's own background-tab scheduling is covered by the
    --disable-*-throttling / --disable-features=CalculateNativeWinOcclusion
    launch flags, but these quick-commerce sites' own React code ALSO reads
    the Page Visibility API (document.hidden / visibilityState) directly
    and pauses its own lazy-loaded/virtualized card rendering when it
    thinks the tab isn't visible -- a decision made in the site's JS, which
    the Chrome-level flags don't touch. Confirmed live: minimizing the
    window still silently broke card lists with all those flags in place
    (an off-screen-position workaround was tried and abandoned -- Windows/
    Chrome can snap an off-screen window back into view, so it didn't
    reliably keep the page "visible" either). This overrides
    document.hidden/visibilityState at the JS level so the page always
    reports itself visible, and swallows the visibilitychange/blur events
    the site would otherwise react to. Installed via
    Page.addScriptToEvaluateOnNewDocument, so it re-applies on every
    driver.get() for the life of this session -- call once per driver.
    """
    script = """
    Object.defineProperty(document, 'hidden', {get: () => false, configurable: true});
    Object.defineProperty(document, 'visibilityState', {get: () => 'visible', configurable: true});
    Object.defineProperty(document, 'webkitHidden', {get: () => false, configurable: true});
    Object.defineProperty(document, 'webkitVisibilityState', {get: () => 'visible', configurable: true});
    ['visibilitychange', 'webkitvisibilitychange', 'blur'].forEach(function(evt) {
        window.addEventListener(evt, function(e) { e.stopImmediatePropagation(); }, true);
        document.addEventListener(evt, function(e) { e.stopImmediatePropagation(); }, true);
    });
    """
    try:
        driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {'source': script})
    except Exception:
        pass


def dismiss_blinkit_interstitials(driver, timeout: float = 5) -> None:
    """Blinkit now shows a "Get the app" nag (DownloadAppModal) on every
    single page load within a session, and a one-time "Select your
    location" chooser (GetLocationModal: "Use my location" / "Select
    manually") before the real location-search input ever exists in the
    DOM. Confirmed live via DOM inspection: neither modal is optional or
    occasional -- the search input this scraper waits for simply is not
    present until both are dismissed, so the previous code's 8s
    WebDriverWait for it timed out with an empty (message-less) Selenium
    TimeoutException on every locality. Clicks go through JS (element.click())
    since these modals animate in and Selenium's native click can reject the
    element as not-yet-interactable mid-animation.
    """
    try:
        close_img = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'img[alt="Close Slider"]'))
        )
        driver.execute_script("arguments[0].parentElement.click();", close_img)
        time.sleep(0.5)
    except Exception:
        pass

    try:
        select_manually = driver.find_element(
            By.XPATH,
            "//div[contains(@class,'SelectManually')]"
        )
        driver.execute_script("arguments[0].click();", select_manually)
        time.sleep(0.5)
    except Exception:
        pass

    # Both modals above are one-time/every-load quirks of a session with no
    # locality confirmed yet. Once the FIRST locality in a run succeeds,
    # every later driver.get() lands straight on the homepage with that
    # locality already showing in the header -- neither modal reappears and
    # the search input never re-enters the DOM at all, which is what made
    # every locality after the first fail all 3 attempts with an empty
    # TimeoutException. Confirmed live via DOM inspection: the header's
    # location bar (class prefix "LocationBar__Container") is the toggle
    # that reopens the same location-search modal to change locality.
    try:
        driver.find_element(By.XPATH, "//input[@placeholder='search delivery location']")
    except Exception:
        try:
            location_bar = driver.find_element(By.CSS_SELECTOR, '[class*="LocationBar__Container"]')
            driver.execute_script("arguments[0].click();", location_bar)
            time.sleep(1)
        except Exception:
            pass


def jittered_sleep(base_s: float, jitter_s: float = 1.0) -> None:
    time.sleep(base_s + random.uniform(0, jitter_s))


def shard_localities(localities: list, shard_index: int, num_shards: int) -> list:
    """Round-robin split so each shard gets an even spread across the whole
    list rather than one contiguous block -- if a particular city has
    connectivity/blocking trouble, that risk is spread across shards
    instead of concentrated in whichever shard owns that city's block."""
    return localities[shard_index::num_shards]


def should_restart_driver(locality_index: int, restart_every: int = 25) -> bool:
    return locality_index > 0 and locality_index % restart_every == 0


def is_dead_session_error(exc: Exception) -> bool:
    msg = str(exc).lower()
    return any(marker in msg for marker in DEAD_SESSION_MARKERS)


class IncrementalWorkbook:
    """One in-memory openpyxl workbook, appended to as rows come in and
    saved periodically -- replaces rewriting the whole xlsx file (or, for
    Zepto's old pattern, round-tripping it through disk) on every save
    point, which gets slower as the run accumulates rows."""

    def __init__(self, path: Path, columns: list[str]):
        self.path = Path(path)
        self.columns = columns
        if self.path.exists():
            self.wb = load_workbook(self.path)
            self.ws = self.wb.active
        else:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.append(columns)

    def append_row(self, row: dict) -> None:
        self.ws.append([row.get(c) for c in self.columns])

    def save(self) -> None:
        self.wb.save(self.path)

    def done_keys(self, key_columns: list[str]) -> set[str]:
        idxs = [self.columns.index(c) for c in key_columns]
        keys = set()
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            keys.add("|".join(str(row[i]) for i in idxs))
        return keys
