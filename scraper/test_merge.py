from openpyxl import Workbook, load_workbook

from _merge import merge_shards


def _write_shard(path, columns, rows):
    wb = Workbook()
    ws = wb.active
    ws.append(columns)
    for row in rows:
        ws.append(row)
    wb.save(path)


def test_merge_shards_combines_and_sorts(tmp_path):
    columns = ["City", "Locality", "Brand Searched", "Price"]
    shard0 = tmp_path / "shard0.xlsx"
    shard1 = tmp_path / "shard1.xlsx"
    _write_shard(shard0, columns, [
        ("Bangalore", "Koramangala", "Quaker", 86),
    ])
    _write_shard(shard1, columns, [
        ("Bangalore", "Indiranagar", "Pintola", 550),
    ])
    output = tmp_path / "combined.xlsx"

    rank = {("Bangalore", "Indiranagar"): 0, ("Bangalore", "Koramangala"): 1}

    def sort_key(row):
        return (rank.get((row["City"], row["Locality"]), 999), row["Brand Searched"])

    n = merge_shards([shard0, shard1], output, columns, sort_key)

    assert n == 2
    result = load_workbook(output)
    rows = list(result.active.iter_rows(values_only=True))
    assert rows[0] == tuple(columns)
    assert rows[1] == ("Bangalore", "Indiranagar", "Pintola", 550)
    assert rows[2] == ("Bangalore", "Koramangala", "Quaker", 86)


def test_merge_shards_skips_missing_shard_files(tmp_path):
    columns = ["City", "Locality"]
    shard0 = tmp_path / "shard0.xlsx"
    missing = tmp_path / "does_not_exist.xlsx"
    _write_shard(shard0, columns, [("Bangalore", "Koramangala")])
    output = tmp_path / "combined.xlsx"

    n = merge_shards([shard0, missing], output, columns, lambda row: row["Locality"])

    assert n == 1


def test_merge_shards_overwrites_existing_output(tmp_path):
    columns = ["City", "Locality"]
    shard0 = tmp_path / "shard0.xlsx"
    _write_shard(shard0, columns, [("Bangalore", "Koramangala")])
    output = tmp_path / "combined.xlsx"
    _write_shard(output, columns, [("Stale", "Data")])  # simulate a prior merge

    n = merge_shards([shard0], output, columns, lambda row: row["Locality"])

    assert n == 1
    result = load_workbook(output)
    rows = list(result.active.iter_rows(values_only=True))
    assert rows == [tuple(columns), ("Bangalore", "Koramangala")]


def test_merge_shards_returns_zero_when_no_shards_exist(tmp_path):
    output = tmp_path / "combined.xlsx"
    n = merge_shards([tmp_path / "missing.xlsx"], output, ["City"], lambda row: row["City"])
    assert n == 0
    assert output.exists()
