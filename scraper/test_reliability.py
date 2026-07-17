import time

import pytest
from openpyxl import load_workbook

from _reliability import (
    IncrementalWorkbook,
    defeat_visibility_throttling,
    is_blocked,
    is_dead_session_error,
    jittered_sleep,
    keep_window_unminimized,
    shard_localities,
    should_restart_driver,
    wait_for_manual_unblock,
)


class FakeDriver:
    def __init__(self, title="GOAT Life Oats", body="normal page content"):
        self.title = title
        self._body = body
        self.cdp_calls = []

    @property
    def page_source(self):
        return self._body

    def execute_cdp_cmd(self, cmd, params):
        self.cdp_calls.append((cmd, params))

    def set_window_rect(self, x=None, y=None, width=None, height=None):
        self.window_rect = (x, y, width, height)


def test_is_blocked_detects_title_keywords():
    assert is_blocked(FakeDriver(title="Please solve this CAPTCHA")) is True
    assert is_blocked(FakeDriver(title="Robot check")) is True
    assert is_blocked(FakeDriver(title="Normal Product Page")) is False


def test_is_blocked_detects_body_markers():
    assert is_blocked(FakeDriver(title="Blinkit", body="<div>Access Denied</div>")) is True
    assert is_blocked(FakeDriver(title="Blinkit", body="AwsWafIntegration challenge")) is True
    assert is_blocked(FakeDriver(title="Blinkit", body="<div>Yoga Bar Oats ₹399</div>")) is False


def test_keep_window_unminimized_shrinks_to_small_onscreen_size():
    driver = FakeDriver()
    keep_window_unminimized(driver)
    x, y, width, height = driver.window_rect
    assert width > 0 and height > 0  # a real, non-zero, non-minimized size


def test_keep_window_unminimized_swallows_errors():
    class BrokenDriver(FakeDriver):
        def set_window_rect(self, **kwargs):
            raise Exception("no window to resize")

    keep_window_unminimized(BrokenDriver())  # must not raise


def test_defeat_visibility_throttling_installs_script_via_cdp():
    driver = FakeDriver()
    defeat_visibility_throttling(driver)
    assert len(driver.cdp_calls) == 1
    cmd, params = driver.cdp_calls[0]
    assert cmd == "Page.addScriptToEvaluateOnNewDocument"
    assert "'hidden'" in params["source"]
    assert "visibilityState" in params["source"]


def test_defeat_visibility_throttling_swallows_cdp_errors():
    class BrokenDriver(FakeDriver):
        def execute_cdp_cmd(self, cmd, params):
            raise Exception("no CDP support")

    defeat_visibility_throttling(BrokenDriver())  # must not raise


def test_is_blocked_detects_zepto_login_wall():
    assert is_blocked(FakeDriver(
        title="Zepto",
        body="<div>Oops! Please login to continue searching</div>",
    )) is True
    # Every normal Zepto page has a "Login" link in the header -- must not
    # false-positive on the bare word.
    assert is_blocked(FakeDriver(
        title="Zepto",
        body="<header><a>Login</a></header><div>True Elements Oats ₹299</div>",
    )) is False


def test_wait_for_manual_unblock_returns_true_immediately_if_not_blocked():
    driver = FakeDriver(title="Normal page")
    beeped = []
    assert wait_for_manual_unblock(driver, beep_fn=lambda: beeped.append(1), poll_s=0.01) is True
    assert beeped == []


def test_wait_for_manual_unblock_clears_after_driver_state_changes():
    driver = FakeDriver(title="CAPTCHA")
    beeped = []

    def unblock_after_beep():
        beeped.append(1)
        driver.title = "Normal page"

    assert wait_for_manual_unblock(driver, beep_fn=unblock_after_beep, poll_s=0.01, max_wait_s=1) is True
    assert beeped == [1]


def test_wait_for_manual_unblock_gives_up_after_max_wait():
    driver = FakeDriver(title="CAPTCHA")
    result = wait_for_manual_unblock(driver, beep_fn=lambda: None, poll_s=0.01, max_wait_s=0.05)
    assert result is False


def test_jittered_sleep_sleeps_at_least_base_duration():
    start = time.monotonic()
    jittered_sleep(0.05, jitter_s=0.05)
    assert time.monotonic() - start >= 0.05


def test_should_restart_driver_fires_every_n_localities():
    assert should_restart_driver(0, restart_every=25) is False
    assert should_restart_driver(24, restart_every=25) is False
    assert should_restart_driver(25, restart_every=25) is True
    assert should_restart_driver(50, restart_every=25) is True
    assert should_restart_driver(26, restart_every=25) is False


def test_is_dead_session_error_matches_known_messages():
    assert is_dead_session_error(Exception("chrome not reachable")) is True
    assert is_dead_session_error(Exception("invalid session id")) is True
    assert is_dead_session_error(Exception("session deleted because of page crash")) is True
    assert is_dead_session_error(Exception("element not found")) is False


def test_incremental_workbook_appends_and_saves(tmp_path):
    path = tmp_path / "out.xlsx"
    wb = IncrementalWorkbook(path, columns=["City", "Locality", "Price"])
    wb.append_row({"City": "Bangalore", "Locality": "Indiranagar", "Price": 99})
    wb.append_row({"City": "Delhi", "Locality": "Saket", "Price": 119})
    wb.save()

    reloaded = load_workbook(path)
    ws = reloaded.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("City", "Locality", "Price")
    assert rows[1] == ("Bangalore", "Indiranagar", 99)
    assert rows[2] == ("Delhi", "Saket", 119)


def test_incremental_workbook_resumes_from_existing_file(tmp_path):
    path = tmp_path / "out.xlsx"
    wb1 = IncrementalWorkbook(path, columns=["City", "Locality", "Price"])
    wb1.append_row({"City": "Bangalore", "Locality": "Indiranagar", "Price": 99})
    wb1.save()

    wb2 = IncrementalWorkbook(path, columns=["City", "Locality", "Price"])
    keys = wb2.done_keys(["City", "Locality"])
    assert keys == {"Bangalore|Indiranagar"}

    wb2.append_row({"City": "Delhi", "Locality": "Saket", "Price": 119})
    wb2.save()

    reloaded = load_workbook(path)
    rows = list(reloaded.active.iter_rows(values_only=True))
    assert len(rows) == 3  # header + 2 data rows


def test_shard_localities_round_robin_covers_all_with_no_duplicates():
    localities = list(range(10))
    shard0 = shard_localities(localities, 0, 3)
    shard1 = shard_localities(localities, 1, 3)
    shard2 = shard_localities(localities, 2, 3)

    assert shard0 == [0, 3, 6, 9]
    assert shard1 == [1, 4, 7]
    assert shard2 == [2, 5, 8]
    assert sorted(shard0 + shard1 + shard2) == localities


def test_shard_localities_single_shard_returns_everything():
    localities = ["a", "b", "c"]
    assert shard_localities(localities, 0, 1) == localities
